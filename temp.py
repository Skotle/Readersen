import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import hashlib

def get_color_from_name(name):
    """이름을 기반으로 고유한 파스텔톤 배경색 생성"""
    if not name or name == "-":
        return "FFFFFF" # 기본 흰색
    
    # 이름 해시값을 이용해 16진수 색상 추출
    hash_obj = hashlib.md5(name.encode())
    hex_color = hash_obj.hexdigest()[:6].upper()
    
    # 너무 어두운 색은 가독성이 떨어지므로 명도를 보정 (간이 보정)
    # 뒤의 2자리를 FF로 고정하거나 하여 밝은 색 계열로 유도 가능
    return hex_color

def save_final_ranking_report():
    # 분석 기준 선택
    mode = input("분석 단위를 선택하세요 (M: 월간, W: 주간, D: 일간): ").upper()
    if mode not in ['M', 'W', 'D']:
        print("잘못된 입력입니다.")
        return

    pattern = re.compile(r"\[(.+?)\]\s\[(.+?)\]\s\[(.+?)\]")
    raw_data = []
    now = datetime.now()
    blacklisted_periods = ['2023-09'] # 이상값 기간

    # 1. 데이터 읽기 및 전처리
    try:
        with open('daily-data.txt', 'r', encoding='utf-8') as f:
            for line in f:
                match = pattern.search(line)
                if match:
                    nick, uid, date_str = match.groups()
                    try:
                        dt = datetime.strptime(date_str.strip(), "%Y-%m-%d %H:%M:%S")
                        p_month = dt.strftime("%Y-%m")
                        
                        # 필터링: 미래 데이터, 블랙리스트, 2024년 이전 데이터 제외
                        if dt > now or p_month in blacklisted_periods or dt.year < 2024:
                            continue
                        
                        # 기간 설정
                        if mode == 'D':   # 일간
                            p = dt.strftime("%Y-%m-%d")
                        elif mode == 'W': # 주간 (해당 주의 월요일 기준)
                            p = (dt.date() - pd.Timedelta(days=dt.weekday())).strftime("%Y-%m-%d(주)")
                        else:             # 월간
                            p = p_month
                        
                        raw_data.append({'Period': p, 'User': nick})
                    except ValueError: continue
    except FileNotFoundError:
        print("daily-data.txt 파일을 찾을 수 없습니다.")
        return

    if not raw_data:
        print("분석할 유효 데이터가 없습니다.")
        return

    df = pd.DataFrame(raw_data)
    
    # 2. 기간별 순수 활동량 집계
    periodic_counts = df.groupby(['Period', 'User']).size().unstack(fill_value=0)
    
    # 3. 순위 계산 (동점자 처리는 먼저 발견된 사람 우선)
    rank_df = periodic_counts.rank(axis=1, method='first', ascending=False)

    # 4. 결과 행 구성 (날짜 + 상위 10명 이름(증가치))
    ranking_rows = []
    for period in rank_df.index:
        day_rank = rank_df.loc[period].sort_values()
        top_10_names = day_rank.head(10).index.tolist()
        
        row = [period]
        for name in top_10_names:
            count = periodic_counts.loc[period, name]
            if count > 0:
                row.append(f"{name}({int(count)})")
            else:
                row.append("-")
        
        while len(row) < 11: row.append("")
        ranking_rows.append(row)

    # 5. 엑셀 저장
    columns = ['Period', '1위', '2위', '3위', '4위', '5위', '6위', '7위', '8위', '9위', '10위']
    final_df = pd.DataFrame(ranking_rows, columns=columns)
    
    mode_name = "monthly" if mode == 'M' else "weekly" if mode == 'W' else "daily"
    output_file = f'rank_report_{mode_name}.xlsx'
    final_df.to_excel(output_file, index=False)

    # 6. 엑셀 색상 및 스타일 입히기
    wb = load_workbook(output_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=11):
        for cell in row:
            if cell.value and cell.value != "-":
                # 이름 추출 (괄호 전까지)
                pure_name = cell.value.split('(')[0]
                bg_color = get_color_from_name(pure_name)
                
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                cell.font = Font(bold=True) # 글씨 굵게

    # 열 너비 자동 맞춤
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    wb.save(output_file)
    print(f"--- 분석 완료 ---")
    print(f"기준: {mode_name}")
    print(f"파일 저장: {output_file}")

# 실행
if __name__ == "__main__":
    save_final_ranking_report()